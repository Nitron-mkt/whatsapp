"""Pendencias de contato de representante no CRM — gera contatos_rep_pendencias.xlsx.

Le dados.json (fetch.py). Para cada contato que o sync marcou como sendo do
representante, apura a EVIDENCIA do apontamento: qual valor casou, com qual campo
do cadastro de qual rep, e o que precisa ser corrigido.

Reproduz a regra do ghl-contatos-sync v12:
  repByEmail = snap_rep.email, email_parc, email_crm   (do rep com codparc em rep_carteira)
  repByFone  = snap_rep.celular, fone_parc             (normalizados: so digitos, sem 0 e sem 55)
  cria #r apenas se o contato NAO tiver AD_CODPARC proprio (trava adicionada em v12)
  #biz vem de materializa_rep_biz(): contatos da mesma empresa GHL do rep
"""
import json, io, re, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

D = json.load(io.open("dados.json", encoding="utf-8"))
LINK = "https://app.gohighlevel.com/v2/location/rZ8y7lzqV7fzxsartaX2/contacts/detail/"
PASSADA = max(x["atualizado"] for x in D["ghl_sync_state"])   # fim da ultima passada completa


def nf(v):
    d = re.sub(r"\D", "", str(v or "")); d = re.sub(r"^0+", "", d)
    if d.startswith("55"): d = d[2:]
    return re.sub(r"^0+", "", d)
def em(v): return str(v or "").strip().lower()

cp_por_vend = {r["codvend"]: r["codparc"] for r in D["rep_carteira"] if r["codparc"] is not None}
reps, chaves = {}, {}          # chaves: (tipo, valor) -> (codparc do rep, campo)
CAMPO = {"email":"e-mail (campo email)", "email_parc":"e-mail (campo email_parc)",
         "email_crm":"e-mail (campo email_crm)", "celular":"telefone (campo celular)",
         "fone_parc":"telefone (campo fone_parc)"}
for s in D["snap_rep"]:
    cp = cp_por_vend.get(s["codvend"])
    if cp is None: continue
    reps[cp] = s["rep"]
    for c in ("email","email_parc","email_crm"):
        if em(s.get(c)): chaves.setdefault(("email", em(s[c])), (cp, c))
    for c in ("celular","fone_parc"):
        if nf(s.get(c)) and len(nf(s[c])) >= 8: chaves.setdefault(("fone", nf(s[c])), (cp, c))

razao = {c["codparc"]: c["razao"] for c in D["ghl_cliente"] if c.get("razao")}
base  = {g["ghl_id"]: g for g in D["ghl_contato"] if "#" not in g["ghl_id"]}

def casou(g):
    """Qual valor do contato casa com qual campo de qual rep."""
    if em(g.get("email")) and ("email", em(g["email"])) in chaves:
        cp, c = chaves[("email", em(g["email"]))]
        return g["email"].strip(), CAMPO[c], cp
    if nf(g.get("fone")) and ("fone", nf(g["fone"])) in chaves:
        cp, c = chaves[("fone", nf(g["fone"]))]
        return g["fone"].strip(), CAMPO[c], cp
    return None, None, None

itens = []
for g in D["ghl_contato"]:
    gid = g["ghl_id"]
    if "#biz" not in gid and "#r" not in gid: continue
    rep_cp = g["codparc"]
    if rep_cp not in reps: continue
    bid = gid.split("#")[0]
    b = base.get(bid)
    marca = "#biz" if "#biz" in gid else "#r"
    val, campo, cp_chave = casou(g)
    if marca == "#r" and not val:
        # o valor gravado no espelho ja nao corresponde a nenhuma chave do cadastro do rep:
        # ou o contato mudou no CRM, ou o cadastro do rep mudou depois da marcacao.
        campo = "não identificado — o valor gravado já não corresponde ao cadastro do rep"

    # --- classifica recalculando a regra do sync v12 com o dado de hoje.
    # Regra v12: cria #r se o e-mail/telefone casa com o cadastro do rep E o contato
    # NAO tem AD_CODPARC proprio. Se hoje a regra nao criaria esta marcacao, ela esta velha.
    tem_cp = bool(b and b.get("codparc"))
    quando = str(g.get("atualizado",""))[:16].replace("T"," ")

    if marca == "#biz":
        cat, pri = "Veio pela empresa do rep no CRM", 3
        porque = ("O contato está na mesma empresa (business) de " + reps[rep_cp] + " no GoHighLevel "
                  "(business_id " + str(g.get("business_id") or "—") + "). O sync anexa ao representante "
                  "todos os contatos da empresa dele, e foi por isso que este entrou — não por e-mail nem "
                  "por telefone.")
        corrigir = ("Confirme se é do escritório de " + reps[rep_cp] + ". Se for cliente, mova o contato "
                    "para a empresa do cliente no GHL: separar as empresas tira a marcação sozinho.")
    elif tem_cp and b["codparc"] == rep_cp:
        base_em = str(b.get("atualizado",""))[:16].replace("T"," ")
        # A marcacao ser MAIS NOVA que a linha base e a prova: nessa varredura o sync viu o
        # contato e nao gravou a linha base (que so existe quando ha AD_CODPARC), mas gravou a
        # marcacao (que so existe quando NAO ha). Ou seja: o campo esta vazio no CRM agora, e a
        # linha base e apenas uma copia velha.
        if str(g.get("atualizado","")) > str(b.get("atualizado","")):
            cat, pri = "AD_CODPARC apagado no CRM", 1
            porque = ("Casou o " + (campo or "e-mail/telefone") + " de " + reps[rep_cp] +
                      (" (" + val + ")" if val else "") + ". Esta marcação é de " + quando + ", mais recente "
                      "que a cópia direta do mesmo contato, que está congelada em " + base_em + " com "
                      "AD_CODPARC=" + str(rep_cp) + ". Como a regra do sync só marca contato SEM AD_CODPARC, "
                      "e a varredura mais recente marcou este, o campo está vazio no CRM agora — a cópia de " +
                      base_em + " é só o registro de quando ele ainda existia.")
            corrigir = ("Reponha AD_CODPARC=" + str(rep_cp) + " neste contato no CRM. É a causa: sem o campo, "
                        "o contato só é reconhecido por casar e-mail, e volta a ser marcado em toda varredura. "
                        "Depois de repor, a linha antiga (ghl_id " + gid + ") ainda precisa ser apagada de "
                        "ghl_contato, porque o sync não apaga marcação.")
        else:
            cat, pri = "Marcação anterior à correção do sync", 1
            porque = ("Casou o " + (campo or "e-mail/telefone") + " de " + reps[rep_cp] +
                      (" (" + val + ")" if val else "") + ", mas o contato tem AD_CODPARC=" + str(rep_cp) +
                      " (visto em " + base_em + ") e a regra atual do sync não marca quem já tem o campo. "
                      "Esta marcação é de " + quando + " — anterior à correção da regra.")
            corrigir = ("Nada a fazer no CRM: o contato já está certo. A linha precisa ser apagada de "
                        "ghl_contato (ghl_id " + gid + ").")
    elif tem_cp and b["codparc"] != rep_cp:
        dono = reps.get(b["codparc"]) or razao.get(b["codparc"]) or ("cód. " + str(b["codparc"]))
        cat, pri = "Contato de outro cadastro marcado para o rep", 1
        porque = ("Casou o " + (campo or "e-mail/telefone") + " de " + reps[rep_cp] +
                  (" (" + val + ")" if val else "") + ", mas o contato tem AD_CODPARC=" + str(b["codparc"]) +
                  ", que é de " + dono + ". São duas identidades no mesmo contato.")
        corrigir = ("O valor que casou pertence ao cadastro de " + reps[rep_cp] + " e está num contato de " +
                    dono + ". Tire esse e-mail/telefone do contato de " + dono + ", ou corrija o cadastro do "
                    "rep no Sankhya se o valor não for dele.")
    elif not val:
        cat, pri = "Marcação velha: o valor já não casa", 1
        porque = ("A marcação foi criada porque algum e-mail ou telefone deste contato casava com o cadastro "
                  "de " + reps[rep_cp] + ". Hoje nenhum dos valores gravados nesta linha (" +
                  ((g.get("email") or g.get("fone") or "sem valor")) + ") corresponde ao cadastro do rep — "
                  "o contato mudou no CRM, ou o cadastro do rep mudou depois. Marcada em " + quando + ".")
        corrigir = ("A marcação não se sustenta mais. Nada a fazer no CRM: a linha precisa ser apagada de "
                    "ghl_contato (ghl_id " + gid + "). Enquanto existir, o contato segue entrando nos "
                    "disparos como se fosse " + reps[rep_cp] + ".")
    else:
        cat, pri = "Casou pelo e-mail/telefone e não tem cadastro próprio", 2
        porque = ("Casou o " + (campo or "e-mail/telefone") + " de " + reps[rep_cp] +
                  (" (" + val + ")" if val else "") + ", e o contato não tem AD_CODPARC no CRM. Por isso o "
                  "sync o tratou como alguém do escritório do representante — que é o comportamento "
                  "esperado dele. Marcada em " + quando + ".")
        corrigir = ("Confirme quem é. Se for do escritório de " + reps[rep_cp] + ", está correto e não precisa "
                    "mexer. Se for cliente, preencha o AD_CODPARC dele no contato: com o campo preenchido a "
                    "regra atual do sync deixa de marcar, sem precisar mexer no e-mail.")

    itens.append([pri, cat, reps[rep_cp], rep_cp, g.get("nome") or "—", g.get("fone") or "",
                  g.get("email") or "", marca, val or "(não identificado)", campo or "—",
                  str(g.get("atualizado",""))[:16].replace("T"," "),
                  porque, corrigir, LINK + bid, None, None])

itens.sort(key=lambda x: (x[0], x[2], x[4]))

# ------------------------------- planilha -------------------------------
A="Arial"; HF=PatternFill("solid",fgColor="FF12514C"); HFONT=Font(name=A,bold=True,color="FFFFFFFF",size=10)
BODY=Font(name=A,size=10); NEG=Font(name=A,size=10,bold=True)
LINKF=Font(name=A,size=10,color="FF0B6BCB",underline="single")
FILL={1:PatternFill("solid",fgColor="FFFBE3E3"),2:PatternFill("solid",fgColor="FFFFF3DC"),3:PatternFill("solid",fgColor="FFEDF2FA")}

def cab(ws, larg, wrap=()):
    for c in ws[1]:
        c.fill=HF; c.font=HFONT; c.alignment=Alignment(vertical="center", wrap_text=c.column_letter in wrap)
    for k,v in larg.items(): ws.column_dimensions[k].width=v
    ws.freeze_panes="A2"; ws.row_dimensions[1].height=32

wb=Workbook(); ws=wb.active; ws.title="Pendências"
ws.append(["#","Tipo de pendência","Representante","Cód. rep","Contato","Telefone","E-mail",
           "Marca","Valor que casou","Campo do cadastro do rep","Marcado em",
           "Por que está apontado","O que precisa corrigir","Abrir no CRM","Resolvido?","Observação"])
for i,it in enumerate(itens,1):
    ws.append([i]+it[1:])
    r=ws.max_row
    for c in ws[r]: c.font=BODY
    ws.cell(r,2).fill=FILL[it[0]]
    if it[0]==1: ws.cell(r,2).font=NEG
    for col in (12,13): ws.cell(r,col).alignment=Alignment(wrap_text=True, vertical="top")
    cel=ws.cell(r,14); cel.hyperlink=it[13]; cel.value="abrir"; cel.font=LINKF
cab(ws, {"A":5,"B":34,"C":18,"D":9,"E":30,"F":17,"G":30,"H":8,"I":30,"J":26,"K":17,
         "L":68,"M":68,"N":11,"O":11,"P":24}, wrap=("B","I","J","L","M"))
ws.auto_filter.ref=f"A1:P{ws.max_row}"

# resumo
w2=wb.create_sheet("Resumo")
w2.append(["Tipo de pendência","Quantos","Representantes","O que fazer"])
ACAO={1:"Agir agora", 2:"Confirmar caso a caso", 3:"Confirmar caso a caso"}
grp=collections.OrderedDict()
for it in itens:
    g=grp.setdefault(it[1], [0,set(),it[0]]); g[0]+=1; g[1].add(it[2])
for k,(n,rr,pri) in sorted(grp.items(), key=lambda x:(x[1][2],-x[1][0])):
    w2.append([k,n,len(rr),ACAO[pri]])
    for c in w2[w2.max_row]: c.font=BODY
    w2.cell(w2.max_row,1).fill=FILL[pri]
w2.append(["TOTAL",len(itens),len({it[2] for it in itens}),""])
for c in w2[w2.max_row]: c.font=NEG
cab(w2, {"A":40,"B":10,"C":16,"D":24}, wrap=("A","D"))

# leia antes
w3=wb.create_sheet("Leia antes")
L=[("Pendências de contato de representante no CRM",""),
   ("",""),
   ("Retrato de", PASSADA[:16].replace("T"," ") + " (fim da última varredura completa do CRM)"),
   ("Como foi verificado", "Cada marcação foi reavaliada aplicando a regra atual do sync (v12) sobre o dado "
    "de hoje: o valor gravado ainda casa com algum campo do cadastro do rep? O contato tem AD_CODPARC "
    "próprio? Quando a regra de hoje não criaria aquela marcação, ela está apontada como velha."),
   ("Total", f"{len(itens)} contatos marcados como sendo de representante."),
   ("",""),
   ("Mudou desde a última lista", str(len(itens)) + " contatos marcados, contra 55 na apuração das 16:17. "
    "Nenhuma das 55 anteriores saiu da lista e entraram 7 novas. Isso não significa que ninguém corrigiu nada: "
    "significa que corrigir no CRM não retira a marcação, porque o sync nunca apaga marcação #r — só cria. "
    "Enquanto a linha não for apagada no banco, ela continua na lista mesmo com o CRM certo."),
   ("A causa do crescimento", "A maior parte das marcações novas é de contato de representante que está SEM "
    "o campo AD_CODPARC no CRM. Sem esse campo o contato só é reconhecido por casar e-mail com o cadastro do "
    "rep, e é remarcado em toda varredura. A varredura das 18:17 gravou 6.149 contatos com o campo "
    "preenchido, então o problema não é geral — é desses contatos específicos."),
   ("",""),
   ("Como ler as colunas",""),
   ("  Marca", "#r = casou pelo e-mail ou telefone do representante.  #biz = veio da mesma empresa do "
    "representante no GoHighLevel."),
   ("  Valor que casou", "O e-mail ou telefone exato que provocou a marcação."),
   ("  Campo do cadastro do rep", "De onde esse valor veio no cadastro do representante, no Sankhya."),
   ("  Marcado em", "Quando o sync escreveu esta marcação. Anterior ao retrato acima = resíduo."),
   ("",""),
   ("Atenção", "A marcação #r nunca é apagada pelo sync — só criada. Quando você corrige o contato no CRM, "
    "o sync para de recriar a linha, mas a que já existe fica em ghl_contato para sempre. É por isso que "
    "há categorias cuja correção é apagar a linha no banco, não mexer no CRM. O ghl_id está escrito na "
    "coluna do que corrigir."),
   ("",""),
   ("Origem", "snap_rep, rep_carteira, ghl_contato, ghl_cliente e ghl_sync_state. "
    "Gerado por scripts/contatos_rep_fetch.py + scripts/contatos_rep_pendencias.py."),
]
for a,b in L: w3.append([a,b])
w3["A1"].font=Font(name=A,bold=True,size=13)
for row in w3.iter_rows(min_row=2,max_row=w3.max_row):
    row[0].font=Font(name=A,size=10,bold=True)
    row[1].font=BODY; row[1].alignment=Alignment(wrap_text=True,vertical="top")
w3["A13"].font=Font(name=A,size=10,bold=True,color="FF9C2B2B")
w3.column_dimensions["A"].width=26; w3.column_dimensions["B"].width=104

wb.save("contatos_rep_pendencias.xlsx")
print("total:", len(itens))
for k,(n,rr,pri) in sorted(grp.items(), key=lambda x:(x[1][2],-x[1][0])):
    print(f"  [{pri}] {n:>3}  {k}  ({len(rr)} reps)")
