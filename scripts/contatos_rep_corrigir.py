"""Gera 'contatos_rep_corrigir.xlsx' — lista de ação para a equipe limpar o CRM.

Le dados.json (produzido por fetch.py) e separa os contatos marcados como do
representante em: erro de fato (contato de outro rep / de cliente / cadastro
duplicado) e situacao esperada (o proprio rep e a equipe do escritorio dele).
"""
import json, io, re, unicodedata, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

D = json.load(io.open("dados.json", encoding="utf-8"))
LOC = "rZ8y7lzqV7fzxsartaX2"
LINK = "https://app.gohighlevel.com/v2/location/" + LOC + "/contacts/detail/"

# universo de reps: quem existe em snap_rep e tem codparc em rep_carteira
cp_por_vend = {r["codvend"]: r["codparc"] for r in D["rep_carteira"] if r["codparc"] is not None}
reps = {}                                    # codparc do rep -> nome
for s in D["snap_rep"]:
    cp = cp_por_vend.get(s["codvend"])
    if cp is not None: reps[cp] = s["rep"]
razao = {c["codparc"]: c["razao"] for c in D["ghl_cliente"] if c.get("razao")}
direto = {g["ghl_id"]: g["codparc"] for g in D["ghl_contato"] if "#" not in g["ghl_id"]}
sync = max(g["atualizado"] for g in D["ghl_contato"] if g.get("atualizado"))

# agrupa as marcacoes por (contato-base, rep)
marc = {}
for g in D["ghl_contato"]:
    gid = g["ghl_id"]
    if "#biz" not in gid and "#r" not in gid: continue
    if not (g.get("fone") or g.get("email")): continue
    cp = g["codparc"]
    if cp not in reps: continue
    k = (gid.split("#")[0], cp)
    m = marc.setdefault(k, {"nome":None,"fone":None,"email":None,"casado":False,"empresa":False})
    for c in ("nome","fone","email"):
        if g.get(c) and not m[c]: m[c] = g[c]
    if "#biz" in gid: m["empresa"] = True
    else: m["casado"] = True

STOP = {"REPRESENTACOES","REPRESENTACAO","REPRESENTANTE","COMERCIAL","COMERCIO","LTDA",
        "EIRELI","ME","CIA","DOS","DAS","DE","DA","DO","E"}
def norm(s):
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii","ignore").decode()
    return re.sub(r"[^A-Z0-9 ]", " ", s.upper())
def parece_o_rep(rep, dono):
    if not dono: return False
    d = norm(dono)
    return any(t in d for t in norm(rep).split() if len(t) >= 4 and t not in STOP)

def porque(m):
    if m["casado"] and m["empresa"]: return "Tem o e-mail/telefone do representante E está na empresa dele no CRM"
    if m["casado"]: return "Tem o e-mail/telefone do representante"
    return "Está na empresa do representante no CRM"

erros, ok = [], []
for (base, cp), m in marc.items():
    rep = reps[cp]
    dono_cp = direto.get(base)
    dono = razao.get(dono_cp) if dono_cp is not None else None
    if dono_cp is None:
        ok.append([rep, m["nome"], m["fone"], m["email"], "Equipe do escritório do representante", LINK+base]); continue
    if dono_cp == cp:
        ok.append([rep, m["nome"], m["fone"], m["email"], "É o próprio representante", LINK+base]); continue
    if dono_cp in reps:
        erros.append([1, "1 · Contato de OUTRO representante", rep, m["nome"], m["fone"], m["email"],
            f'{reps[dono_cp]} (cód. {dono_cp})',
            porque(m),
            f'Este contato é do representante {reps[dono_cp]}. Tire o e-mail/telefone de {rep} deste contato no CRM — hoje ele recebe mensagem dos dois.',
            LINK+base]); continue
    if parece_o_rep(rep, dono):
        erros.append([3, "3 · Cadastro com o nome do rep — verificar", rep, m["nome"], m["fone"], m["email"],
            (dono or "sem razão social") + f' (cód. {dono_cp})',
            porque(m),
            f'O cadastro de destino tem o nome de {rep}. Pode ser o próprio rep cadastrado duas vezes (cód. {dono_cp} e {cp}) ou alguém da família/escritório. Confirme com o comercial: se for o mesmo rep, unifique o cadastro (mexer no CRM não resolve); se for cliente, corrija no CRM.',
            LINK+base]); continue
    erros.append([2, "2 · Contato de CLIENTE na ficha do rep", rep, m["nome"], m["fone"], m["email"],
        (dono or "sem razão social") + f' (cód. {dono_cp})',
        porque(m),
        f'Contato do cliente {dono or dono_cp}. Remova o e-mail/telefone do representante deste contato e, se ele estiver na empresa de {rep} no CRM, mova para a empresa do cliente.',
        LINK+base])

erros.sort(key=lambda r: (r[0], r[2], r[3] or ""))
ok.sort(key=lambda r: (r[0], r[1] or ""))

ARIAL = "Arial"
HF = PatternFill("solid", fgColor="FF12514C")
HFONT = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=10)
BODY = Font(name=ARIAL, size=10)
LINKF = Font(name=ARIAL, size=10, color="FF0B6BCB", underline="single")
FILLS = {1: PatternFill("solid", fgColor="FFFBE3E3"),
         2: PatternFill("solid", fgColor="FFFFF3DC"),
         3: PatternFill("solid", fgColor="FFEDF2FA")}
def cabeca(ws, largs, wrap=()):
    for c in ws[1]:
        c.fill = HF; c.font = HFONT
        c.alignment = Alignment(vertical="center", wrap_text=c.column_letter in wrap)
    for col, w in largs.items(): ws.column_dimensions[col].width = w
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30

wb = Workbook()

# ---- ABA 1: o que corrigir ----
ws = wb.active; ws.title = "Corrigir"
ws.append(["#","O que é","Está na ficha de","Contato","Telefone","E-mail",
           "Na verdade é de","Por que entrou na ficha do rep","O que fazer","Abrir no CRM","Feito?","Observação"])
for i, e in enumerate(erros, 1):
    ws.append([i] + e[1:] + [None, None])
    r = ws.max_row
    for c in ws[r]: c.font = BODY
    ws.cell(r, 2).fill = FILLS[e[0]]
    cel = ws.cell(r, 10); cel.hyperlink = e[9]; cel.value = "abrir"; cel.font = LINKF
    ws.cell(r, 9).alignment = Alignment(wrap_text=True, vertical="top")
    ws.cell(r, 8).alignment = Alignment(wrap_text=True, vertical="top")
cabeca(ws, {"A":5,"B":34,"C":18,"D":30,"E":17,"F":30,"G":34,"H":34,"I":62,"J":11,"K":9,"L":26},
       wrap=("B","G","H","I"))
ws.auto_filter.ref = f"A1:L{ws.max_row}"

# ---- ABA 2: conferido, nao mexer ----
w2 = wb.create_sheet("Não é erro (conferido)")
w2.append(["Representante","Contato","Telefone","E-mail","Por que aparece marcado","Abrir no CRM"])
for o in ok:
    w2.append(o[:5] + [None])
    r = w2.max_row
    for c in w2[r]: c.font = BODY
    cel = w2.cell(r, 6); cel.hyperlink = o[5]; cel.value = "abrir"; cel.font = LINKF
cabeca(w2, {"A":18,"B":34,"C":17,"D":32,"E":38,"F":11}, wrap=("E",))
w2.auto_filter.ref = f"A1:F{w2.max_row}"

# ---- ABA 3: resumo por rep ----
cont = collections.Counter(e[2] for e in erros)
w3 = wb.create_sheet("Por representante")
w3.append(["Representante","A corrigir","De outro rep","De cliente","Nome do rep"])
for rep, n in sorted(cont.items(), key=lambda x: (-x[1], x[0])):
    w3.append([rep, n,
               sum(1 for e in erros if e[2]==rep and e[0]==1),
               sum(1 for e in erros if e[2]==rep and e[0]==2),
               sum(1 for e in erros if e[2]==rep and e[0]==3)])
    for c in w3[w3.max_row]: c.font = BODY
tot = w3.max_row + 1
w3.append(["TOTAL", len(erros),
           sum(1 for e in erros if e[0]==1), sum(1 for e in erros if e[0]==2), sum(1 for e in erros if e[0]==3)])
for c in w3[tot]: c.font = Font(name=ARIAL, size=10, bold=True)
cabeca(w3, {"A":20,"B":11,"C":13,"D":12,"E":18}, wrap=("B","C","D","E"))

# ---- ABA 4: leia antes ----
w4 = wb.create_sheet("Leia antes")
L = [
 ("Contatos de representante para corrigir no CRM",""),
 ("",""),
 ("Retrato de", sync[:16] + " (UTC)"),
 ("O que fazer", f'Abra a aba "Corrigir": {len(erros)} contatos, do mais grave para o menos. '
                 'Cada linha tem o link direto do contato no CRM, o motivo e a ação. '
                 'Marque "Feito?" conforme for resolvendo.'),
 ("O que NÃO fazer", f'A aba "Não é erro (conferido)" tem {len(ok)} contatos que aparecem marcados mas estão certos '
                     '— são o próprio representante e a equipe do escritório dele. Não mexa neles.'),
 ("",""),
 ("Os 3 tipos",""),
 ("  1 · De outro rep", "O contato pertence a outro representante. É o mais grave: os dois recebem a mesma mensagem. "
                        "Corrige no CRM, tirando o e-mail/telefone do rep errado."),
 ("  2 · De cliente", "É contato de um cliente que entrou na ficha do rep. O cliente recebe mensagem escrita para "
                      "representante. Corrige no CRM."),
 ("  3 · Nome do rep", "O cadastro de destino tem o nome do representante — pode ser ele cadastrado duas vezes, "
                       "ou alguém da família/escritório. Confirme com o comercial antes de mexer: se for o mesmo "
                       "rep, o caminho é unificar o cadastro, não editar o CRM."),
 ("",""),
 ("IMPORTANTE", "Limpar no CRM hoje NÃO faz este número cair, e não é falha da equipe. O sistema copia os "
                "contatos do CRM para uma tabela interna e, nessa cópia, a marcação de \"contato do representante\" "
                "nunca é apagada — só criada. Então o contato corrigido no CRM continua marcado aqui. "
                "Pior: essa marca velha faz o sistema puxar de novo todos os contatos da empresa do rep. "
                "Enquanto isso não for ajustado no código, a lista só cresce, por mais que a equipe limpe."),
 ("",""),
 ("Sobre os links", "O link abre o contato no GoHighLevel. Confira o primeiro para garantir que abre na conta certa."),
]
for a, b in L: w4.append([a, b])
w4["A1"].font = Font(name=ARIAL, bold=True, size=13)
for row in w4.iter_rows(min_row=2, max_row=w4.max_row):
    row[0].font = Font(name=ARIAL, size=10, bold=True)
    row[1].font = BODY; row[1].alignment = Alignment(wrap_text=True, vertical="top")
w4["A12"].font = Font(name=ARIAL, size=10, bold=True, color="FF9C2B2B")
w4.column_dimensions["A"].width = 26; w4.column_dimensions["B"].width = 105

wb.save("contatos_rep_corrigir.xlsx")
print(f"corrigir: {len(erros)}  (outro rep {sum(1 for e in erros if e[0]==1)}, "
      f"cliente {sum(1 for e in erros if e[0]==2)}, nome do rep {sum(1 for e in erros if e[0]==3)})")
print(f"nao e erro: {len(ok)}   total marcado: {len(marc)}   reps na lista: {len(cont)}")
