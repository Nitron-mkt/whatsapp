"""Rastreio de contatos dos representantes — gera rastreio_contatos_representantes.xlsx.

Audita, contato por contato, o que o sistema considera ser telefone/e-mail de
representante, e de onde cada um veio. O alvo são contatos de CLIENTE que
entraram na ficha do rep por dedução, porque eles recebem mensagem destinada a
representante na régua de envio.

Universo: reps presentes em snap_rep com codparc em rep_carteira.

Origens e risco:
  OK       cadastro do rep no Sankhya (celular, fone_parc, email, email_parc,
           email_crm), contato do parceiro do rep (snap_contato pelo codparc),
           contato direto no GHL (ghl_contato sem marca no ghl_id) e
           rep_contato_extra (adicionado na tela)
  REVISAR  ghl_contato com marca no ghl_id:
             #biz<codparc>  mesma EMPRESA do rep no GHL
             #r<codparc>    CASADO pelo e-mail/telefone do rep
           A ordem da classificação segue campanhas-preview: #biz antes de #r.

Entrada: rows2.json, com {"sync": <max(ghl_contato.atualizado)>, "rows": [...]},
produzido pela consulta em rastreio_contatos_rep.sql.

O resumo por rep usa valores calculados aqui, não fórmulas: o LibreOffice do
ambiente de geração não abre xlsx, então fórmulas ficariam sem valor em cache e
apareceriam vazias em qualquer leitor que não recalcule.
"""
import json, io, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

_p = json.load(io.open("rows2.json", encoding="utf-8"))
data, SYNC = _p["rows"], _p["sync"]

MSG_CASADO  = "Casou pelo e-mail/telefone do rep. Se for contato de CLIENTE, remova o e-mail/telefone do rep desse contato no GHL."
MSG_EMPRESA = "Veio da mesma EMPRESA do rep no GHL. Se for contato de CLIENTE, separe as empresas no GHL (rep vs cliente)."
def acao(o):
    if o == "CRM: CASADO por email/fone do rep": return MSG_CASADO
    if o == "CRM: via EMPRESA do GHL":           return MSG_EMPRESA
    return None

ARIAL   = "Arial"
HDRFILL = PatternFill("solid", fgColor="FF12514C")
HDRFONT = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=10)
BODY    = Font(name=ARIAL, size=10)
REVFONT = Font(name=ARIAL, size=10, bold=True, color="FF9C2B2B")
REVFILL = PatternFill("solid", fgColor="FFFDECEC")

wb = Workbook()

# ---------- ABA 1: rastreio detalhado ----------
ws = wb.active
ws.title = "Rastreio contatos rep"
H1 = ["Cod Rep","Representante","Codparc do Rep","Canal","Contato (tel/email)","Base",
      "Origem (de onde veio)","Risco","Nome no contato","Empresa GHL (business id)","O que verificar/corrigir"]
ws.append(H1)
for r in data:
    ws.append([str(r["codvend"]), r["rep"], str(r["codparc"]), r["canal"], r["contato"],
               r["base"], r["origem"], r["risco"], r["nome_ct"], r["biz"], acao(r["origem"])])

for c in ws[1]:
    c.fill = HDRFILL; c.font = HDRFONT; c.alignment = Alignment(vertical="center")
for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
    for c in row: c.font = BODY
    rc = row[7]
    if rc.value == "REVISAR":
        rc.font = REVFONT; rc.fill = REVFILL

W1 = {"A":8,"B":22,"C":14,"D":10,"E":34,"F":10,"G":36,"H":10,"I":26,"J":26,"K":62}
for k,v in W1.items(): ws.column_dimensions[k].width = v
ws.freeze_panes = "A2"
ws.auto_filter.ref = f"A1:K{ws.max_row}"
DET = ws.max_row

# ---------- ABA 2: reps a corrigir ----------
por = collections.OrderedDict()
for r in data:
    k = r["codvend"]
    d = por.setdefault(k, {"rep": r["rep"], "codparc": r["codparc"], "rev": 0})
    if r["risco"] == "REVISAR": d["rev"] += 1
alvo = sorted([(k,v) for k,v in por.items() if v["rev"] > 0],
              key=lambda x: (-x[1]["rev"], x[1]["rep"]))
limpos = [v["rep"] for k,v in por.items() if v["rev"] == 0]

w2 = wb.create_sheet("Reps a corrigir")
H2 = ["Cod Rep","Representante","Codparc do Rep","Contatos rastreados","A revisar",
      "Casados por email/fone","Via empresa do GHL","Contatos OK"]
w2.append(H2)
D = "'Rastreio contatos rep'!"
A, G, Hc = f"{D}$A$2:$A${DET}", f"{D}$G$2:$G${DET}", f"{D}$H$2:$H${DET}"
# valores calculados em Python: o LibreOffice deste ambiente nao carrega xlsx,
# entao formulas ficariam sem valor em cache (em branco em qualquer previewer).
agg = {}
for r in data:
    d = agg.setdefault(r["codvend"], {"n":0,"rev":0,"cas":0,"emp":0,"ok":0})
    d["n"] += 1
    if r["risco"] == "REVISAR": d["rev"] += 1
    else: d["ok"] += 1
    if r["origem"] == "CRM: CASADO por email/fone do rep": d["cas"] += 1
    elif r["origem"] == "CRM: via EMPRESA do GHL": d["emp"] += 1
col = [0]*5
for cv, v in alvo:
    a = agg[cv]
    vals = [a["n"], a["rev"], a["cas"], a["emp"], a["ok"]]
    col = [x+y for x,y in zip(col, vals)]
    w2.append([str(cv), v["rep"], str(v["codparc"])] + vals)
tot = w2.max_row + 1
w2.append(["","","TOTAL"] + col)

for c in w2[1]:
    c.fill = HDRFILL; c.font = HDRFONT; c.alignment = Alignment(vertical="center", wrap_text=True)
for row in w2.iter_rows(min_row=2, max_row=w2.max_row):
    for c in row: c.font = BODY
for c in w2[tot]: c.font = Font(name=ARIAL, size=10, bold=True)
for k,v in {"A":8,"B":22,"C":14,"D":13,"E":10,"F":14,"G":14,"H":11}.items():
    w2.column_dimensions[k].width = v
w2.freeze_panes = "A2"
w2.auto_filter.ref = f"A1:H{tot-1}"

# ---------- ABA 3: como ler ----------
w3 = wb.create_sheet("Como ler")
linhas = [
 ("Rastreio de contatos dos representantes", ""),
 ("", ""),
 ("O que é", "Cada linha é UM contato (telefone ou e-mail) que o sistema hoje considera como sendo do representante, "
             "com a origem de onde ele veio. Serve para achar contato de CLIENTE que entrou por engano na ficha do rep."),
 ("Universo", f"{len(por)} representantes — os que existem em snap_rep e têm codparc em rep_carteira."),
 ("Total de linhas", f"{len(data)} contatos rastreados."),
 ("Retrato de", f"ghl_contato sincronizado em {SYNC[:16]} (UTC). O CRM sincroniza sozinho, "
                f"entao os numeros mudam entre apuracoes — esta planilha e um retrato desse instante."),
 ("", ""),
 ("Risco = REVISAR", "Vínculo fraco: o contato foi atribuído ao rep por dedução, não por cadastro. São os dois casos abaixo."),
 ("  CASADO por email/fone", "O contato no GHL tem o mesmo e-mail ou telefone do rep (marca #r no ghl_id). "
                             "Se for contato de CLIENTE, remova o e-mail/telefone do rep desse contato no GHL."),
 ("  Via EMPRESA do GHL", "O contato está na mesma empresa (business) do rep no GHL (marca #biz no ghl_id). "
                          "Se for contato de CLIENTE, separe as empresas no GHL (rep vs cliente)."),
 ("Risco = OK", "Vínculo forte: cadastro do rep no Sankhya, contato do parceiro do rep, "
                "contato direto no GHL (AD_CODPARC do rep) ou adicionado na tela."),
 ("", ""),
 ("Por que importa", "Contato marcado como do rep entra na régua de envio como se fosse o rep. "
                     "Um cliente nessa lista recebe mensagem destinada a representante."),
 ("", ""),
 ("Sem nenhum apontamento", ", ".join(sorted(limpos)) if limpos else "nenhum rep está limpo"),
]
for a,b in linhas: w3.append([a,b])
w3["A1"].font = Font(name=ARIAL, bold=True, size=13)
for row in w3.iter_rows(min_row=2, max_row=w3.max_row):
    row[0].font = Font(name=ARIAL, size=10, bold=True)
    if len(row) > 1:
        row[1].font = BODY; row[1].alignment = Alignment(wrap_text=True, vertical="top")
w3.column_dimensions["A"].width = 26
w3.column_dimensions["B"].width = 104

wb.save("rastreio_contatos_representantes.xlsx")
print("gerado. detalhe:", DET-1, "linhas | reps a corrigir:", len(alvo), "| limpos:", len(limpos), limpos)
