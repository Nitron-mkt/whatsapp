"""Monta 'contatos_rep_conflitos.xlsx' a partir de achados.json (colisoes.py)."""
import json, io, collections
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

P = json.load(io.open("achados.json", encoding="utf-8"))
A, DEST, ORD = P["achados"], P["destinos"], {int(k): v for k, v in P["ordem"].items()}

ARIAL = "Arial"
HF    = PatternFill("solid", fgColor="FF12514C")
HFONT = Font(name=ARIAL, bold=True, color="FFFFFFFF", size=10)
BODY  = Font(name=ARIAL, size=10)
NEG   = Font(name=ARIAL, size=10, bold=True)
FILL  = {1: PatternFill("solid", fgColor="FFFBE3E3"),
         2: PatternFill("solid", fgColor="FFFFF3DC"),
         3: PatternFill("solid", fgColor="FFEDF2FA")}
ST    = {"OK":               PatternFill("solid", fgColor="FFE6F4EA"),
         "NÃO RECEBE NADA":  PatternFill("solid", fgColor="FFFBE3E3")}

def cabeca(ws, largs, wrap=()):
    for c in ws[1]:
        c.fill = HF; c.font = HFONT
        c.alignment = Alignment(vertical="center", wrap_text=c.column_letter in wrap)
    for k, v in largs.items(): ws.column_dimensions[k].width = v
    ws.freeze_panes = "A2"; ws.row_dimensions[1].height = 30

wb = Workbook()

# ---- ABA 1: destino real do disparo ----
ws = wb.active; ws.title = "Destino do disparo"
ws.append(["Representante","WhatsApp que vai ser usado","E-mail que vai ser usado",
           "Instância","Situação","Conflito"])
for d in DEST:
    ws.append(d)
    r = ws.max_row
    for c in ws[r]: c.font = BODY
    cel = ws.cell(r, 5)
    if d[4] == "OK": cel.fill = ST["OK"]
    elif "NÃO RECEBE" in d[4]: cel.fill = ST["NÃO RECEBE NADA"]; cel.font = NEG
    else: cel.fill = FILL[2]
    ws.cell(r, 6).alignment = Alignment(wrap_text=True, vertical="top")
cabeca(ws, {"A":20,"B":26,"C":34,"D":13,"E":32,"F":40}, wrap=("B","C","E","F"))
ws.auto_filter.ref = f"A1:F{ws.max_row}"

# ---- ABA 2: problemas ----
w2 = wb.create_sheet("Problemas")
w2.append(["#","Gravidade","Problema","Canal","Contato","Representante(s)",
           "Detalhe","O que acontece no disparo","O que fazer","Resolvido?"])
for i, a in enumerate(A, 1):
    w2.append([i, ORD[a[0]], a[1], a[2], a[3], a[4], a[5], a[6], a[7], None])
    r = w2.max_row
    for c in w2[r]: c.font = BODY
    w2.cell(r, 2).fill = FILL[a[0]]
    if a[0] == 1: w2.cell(r, 2).font = NEG
    for col in (7, 8, 9): w2.cell(r, col).alignment = Alignment(wrap_text=True, vertical="top")
cabeca(w2, {"A":5,"B":24,"C":40,"D":10,"E":34,"F":26,"G":48,"H":46,"I":48,"J":11},
       wrap=("B","C","G","H","I"))
w2.auto_filter.ref = f"A1:J{w2.max_row}"

# ---- ABA 3: resumo ----
w3 = wb.create_sheet("Resumo")
w3.append(["Problema","Gravidade","Quantos","Representantes atingidos"])
grp = collections.defaultdict(lambda: [0, set(), 0])
for a in A:
    g = grp[a[1]]; g[0] += 1; g[2] = a[0]
    for nm in a[4].split(" | "): g[1].add(nm)
for t, (n, reps_, g) in sorted(grp.items(), key=lambda x: (x[1][2], -x[1][0])):
    w3.append([t, ORD[g], n, len(reps_)])
    for c in w3[w3.max_row]: c.font = BODY
    w3.cell(w3.max_row, 2).fill = FILL[g]
w3.append(["TOTAL","", len(A), len({nm for a in A for nm in a[4].split(" | ")})])
for c in w3[w3.max_row]: c.font = NEG
cabeca(w3, {"A":46,"B":24,"C":10,"D":24}, wrap=("A","B","D"))

# ---- ABA 4: leia antes ----
w4 = wb.create_sheet("Leia antes")
n1 = sum(1 for a in A if a[0] == 1)
nrec = sum(1 for d in DEST if "NÃO RECEBE" in d[4])
socel = sum(1 for d in DEST if d[1] == "—")
L = [
 ("Conflitos de contato antes do disparo aos representantes",""),
 ("",""),
 ("Universo", f'{P["tot_reps"]} representantes · {P["tot_contatos"]} contatos distintos considerados.'),
 ("",""),
 ('Aba "Destino do disparo"', 'O que o disparo do Roteiro de visitas vai realmente usar em cada rep: '
   'o primeiro celular válido e o primeiro e-mail do cadastro, na mesma ordem que o sistema usa. '
   'É a aba para conferir antes de apertar o botão.'),
 ('Aba "Problemas"', f'{len(A)} apontamentos, do mais grave ao menos. Filtre pela coluna Gravidade.'),
 ("",""),
 ("As 3 gravidades",""),
 ("  1 · Erra o destinatário", f'{n1} casos. Ou a mensagem vai para a pessoa errada, ou o rep não recebe. '
   'Resolver antes do disparo.'),
 ("  2 · Pode errar", 'O contato do rep também está em cadastro de cliente. Precisa alguém confirmar de quem é.'),
 ("  3 · Conserte depois", 'Não muda o resultado do disparo: número que o sistema já descarta sozinho, '
   'ou o mesmo rep cadastrado em dois códigos (a mensagem chega na pessoa certa de todo jeito).'),
 ("",""),
 ("Regra do celular", 'O disparo só manda WhatsApp para número com 11 dígitos (DDD + 9 dígitos). '
   'Fixo (10 dígitos) e número curto são descartados em silêncio. '
   f'Hoje {socel} reps não têm celular válido no cadastro e {nrec} não têm nem celular nem e-mail.'),
 ("",""),
 ("Atenção: bases diferentes", 'O Roteiro de visitas usa só o cadastro do rep (celular, fone_parc, email, '
   'email_crm). As outras campanhas de rep usam também os contatos herdados do parceiro do rep, no Sankhya '
   'e no CRM. Por isso um rep pode receber WhatsApp numa campanha e não na outra — está sinalizado.'),
 ("",""),
 ("Origem dos dados", 'snap_rep, rep_carteira, rep_contato_extra, snap_contato e ghl_contato. '
   'Gerado por scripts/contatos_rep_fetch.py + scripts/contatos_rep_colisoes.py.'),
]
for a, b in L: w4.append([a, b])
w4["A1"].font = Font(name=ARIAL, bold=True, size=13)
for row in w4.iter_rows(min_row=2, max_row=w4.max_row):
    row[0].font = Font(name=ARIAL, size=10, bold=True)
    row[1].font = BODY; row[1].alignment = Alignment(wrap_text=True, vertical="top")
w4["A9"].font = Font(name=ARIAL, size=10, bold=True, color="FF9C2B2B")
w4.column_dimensions["A"].width = 26; w4.column_dimensions["B"].width = 104

wb.save("contatos_rep_conflitos.xlsx")
print("gerado |", len(A), "apontamentos |", len(DEST), "reps na aba de destino")
